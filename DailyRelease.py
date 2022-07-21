import os
from datetime import datetime
import pandas as pd
import openpyxl
today=datetime.today().strftime('%Y-%m-%d')

def deleteDN(DNtoDel):
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl02nNav.vbs")
    for DN in DNtoDel:
        os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl02n_DelOneDN.vbs {}".format(DN))
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\exist.vbs")

def releaseRaw():
    print("==========================================")
    today=datetime.today().strftime('%Y-%m-%d')
    filename = today+"-raw.txt"
    path = "C:\\Users\\roshan.liu\\Scripts\\DATA\\SAP_OUTPUT"
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\ReleaseRaw.vbs {} {}".format(path,filename))
    print("==========================================")
    print("Raw Released")
    print("==========================================")
    os.system("wmic printer where name='KONICAMINOLTA-bizhub-C558-B9-F9-11' call setdefaultprinter")


def releaseChecked(SAPtoRel):
    today=datetime.today().strftime('%Y-%m-%d')
    path = "C:\\Users\\roshan.liu\\Scripts\\DATA\\SAP_OUTPUT"
    filename = today+"-solved.txt"
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl10dNav.vbs")
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl10d_InputMultiple.vbs")
    for SAP in SAPtoRel:
        os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl10d_InputMultiple_logOne.vbs {}".format(SAP))
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl10d_InputMultiple_Submit.vbs")
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl10d_Submit.vbs")
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl10d_Submit_SelectAll.vbs")
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl10d_Submit_Release.vbs")
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\vl10d_Submit_Export.vbs {} {}".format(path,filename))
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\exist.vbs")
    os.system("Scripts\\SAP_DailyRelease\\SAP_Scripts\\exist.vbs")
    print("Checked Released")


def refine(TXT):
    with open(TXT,'r') as o:
        lines=o.readlines()
    with open(TXT,'w') as oo:
        for number, line in enumerate(lines):
            if number not in [0,1,2,3]:
                oo.write(line)
def findPartial():
    today=datetime.today().strftime('%Y-%m-%d')
    TXT = r'C:/Users/roshan.liu/Scripts/DATA/SAP_OUTPUT/{}-raw.txt'.format(today)
    DNtoDel = []
    SAPtoRel = []
    target = r'C:/Users/roshan.liu/Scripts/DATA/SAP_OUTPUT/provicional.xlsx'
    df= pd.read_csv(TXT,delim_whitespace=True)
    df.to_excel(target,header=False)
    wb=openpyxl.load_workbook(target)
    ws=wb.active
    count=0
    for row in ws:
        if not all([cell.value == None for cell in row]):
            count += 1
    for row in range(1,count):
        if ws.cell(row,5).value =="2350":
            r=row+1
            partCal=0
            relCal=0
            while ((ws.cell(r,5).value!="2350") & (ws.cell(r,5).value!=None)):
                if len(ws.cell(r,5).value) == 10:
                    relCal+=1
                else:
                    partCal+=1
                r+=1
            if relCal ==partCal:
                SAPtoRel.append(ws.cell(row,1).value)
    for row2 in range(1,count):
        if len(ws.cell(row2,5).value)==10:
            if ws.cell(row2,5).value not in DNtoDel:
                DNtoDel.append(ws.cell(row2,5).value)
    wb2=openpyxl.Workbook()
    ws2=wb2.active
    for r in range(1,len(SAPtoRel)):
        ws2.cell(r,1).value= SAPtoRel[r]
    for rr in range(1,len(DNtoDel)):
        ws2.cell(rr,3).value=DNtoDel[rr]
    wb2.save(r"C:/Users/roshan.liu/Scripts/DATA/SAP_OUTPUT/{}-log.xlsx".format(today))
    wb2.close()
    wb.close()
    os.remove(target)
    deleteDN(DNtoDel)
    print("Partial delivery note deleted")
    return SAPtoRel





releaseRaw()
input("did you throw away?")
releaseChecked(findPartial())
print("Finished")


